"""
============================================================================
 backend/svensson.py
 ---------------------------------------------------------------------------
 Svensson (1994) Yield-Curve Engine.

 Funktionen:
   - zero_rate(maturity, params)        Zero-Rate y(τ) [Prozent]
   - discount_factor(maturity, params)  D(τ) = exp(-r·τ), r in Dezimal
   - params_from_row(row)               DataFrame-Adapter
   - historical_curve(date, df)         Lookup mit ffill (default)
   - shift_curve(params, ...)           Additive β-Shocks
   - curve_grid(params, maturities)     Komplette Kurve über Grid

 Hinweis zur Vola/Korrelationen:
   compute_historical_stats() liegt in 04_fetch_bundesbank_svensson.py
   und wird von dort genutzt — KEINE Duplizierung hier.

 Validierung:
   gegen Sheet 'Risk Free Rates' in Risk_free_rates_stresstest.xlsx.
   Run: python backend/svensson.py
============================================================================
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import NamedTuple, Union

import numpy as np
import pandas as pd

# Pfad-Setup, damit `from config import ...` aus dem Subordner funktioniert
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))


# ----------------------------------------------------------------------
# 1. Datentypen
# ----------------------------------------------------------------------
class SvenssonParams(NamedTuple):
    """Container für die sechs Svensson-Parameter eines Tages.

    β-Werte in Prozentpunkten (Bundesbank-Konvention).
    τ-Werte in Jahren.
    """
    beta0: float
    beta1: float
    beta2: float
    beta3: float
    tau1: float
    tau2: float


ArrayLike = Union[float, np.ndarray, pd.Series]


# ----------------------------------------------------------------------
# 2. Hilfsfunktionen
# ----------------------------------------------------------------------
def _phi(x: np.ndarray) -> np.ndarray:
    """φ(x) = (1 − exp(−x)) / x mit Grenzwert φ(0) = 1.

    Für sehr kleine |x| Taylor-Reihe verwenden, sonst expm1
    für numerische Stabilität.
    """
    x = np.asarray(x, dtype=float)
    with np.errstate(divide="ignore", invalid="ignore"):
        big = -np.expm1(-x) / x
    taylor = 1.0 - 0.5 * x + (x ** 2) / 6.0
    return np.where(np.abs(x) > 1e-8, big, taylor)


# ----------------------------------------------------------------------
# 3. Kern-Mathematik
# ----------------------------------------------------------------------
def zero_rate(
    maturity: ArrayLike,
    params: SvenssonParams,
    *,
    as_decimal: bool = False,
) -> Union[float, np.ndarray]:
    """Svensson-Zero-Rate y(τ) bei Maturity τ (Jahre).

    Formel (Svensson 1994):

        y(τ) = β₀
             + β₁ · φ(τ/τ₁)
             + β₂ · (φ(τ/τ₁) − exp(−τ/τ₁))
             + β₃ · (φ(τ/τ₂) − exp(−τ/τ₂))

    mit φ(x) = (1 − exp(−x)) / x.

    Parameters
    ----------
    maturity : float oder ndarray
        Maturity in Jahren (> 0).
    params : SvenssonParams
        β in Prozentpunkten, τ in Jahren.
    as_decimal : bool, default False
        False → Prozent (Bundesbank/Excel-Konvention, fürs Frontend).
        True  → Dezimal (für Diskontierung, Merton, MC).

    Returns
    -------
    float wenn maturity skalar, ndarray sonst.
    """
    tau = np.asarray(maturity, dtype=float)
    with np.errstate(divide="ignore", invalid="ignore", over="ignore"):
        x1 = tau / params.tau1
        x2 = tau / params.tau2

        phi1 = _phi(x1)
        phi2 = _phi(x2)
        e1 = np.exp(-x1)
        e2 = np.exp(-x2)

        y = (
            params.beta0
            + params.beta1 * phi1
            + params.beta2 * (phi1 - e1)
            + params.beta3 * (phi2 - e2)
        )

    if as_decimal:
        y = y / 100.0

    return float(y) if tau.ndim == 0 else y


def discount_factor(
    maturity: ArrayLike,
    params: SvenssonParams,
) -> Union[float, np.ndarray]:
    """Stetig verzinster Diskontfaktor D(τ) = exp(−r(τ)·τ).

    r wird intern in Dezimal konvertiert; Output ist dimensionslos.
    """
    tau = np.asarray(maturity, dtype=float)
    r = zero_rate(tau, params, as_decimal=True)
    df = np.exp(-r * tau)
    return float(df) if tau.ndim == 0 else df


# ----------------------------------------------------------------------
# 4. DataFrame-Adapter
# ----------------------------------------------------------------------
def params_from_row(row: pd.Series) -> SvenssonParams:
    """Konvertiert eine Zeile des Bundesbank-DataFrames zu SvenssonParams.

    Erwartet die Spalten Beta0, Beta1, Beta2, Beta3, Tau1, Tau2
    (siehe config.BBK_COLUMNS).
    """
    return SvenssonParams(
        beta0=float(row["Beta0"]),
        beta1=float(row["Beta1"]),
        beta2=float(row["Beta2"]),
        beta3=float(row["Beta3"]),
        tau1=float(row["Tau1"]),
        tau2=float(row["Tau2"]),
    )


def historical_curve(
    date: Union[str, pd.Timestamp],
    svensson_df: pd.DataFrame,
    *,
    method: str = "ffill",
) -> SvenssonParams:
    """Lookup historischer Svensson-Parameter zum gegebenen Datum.

    Parameters
    ----------
    date : str oder Timestamp
        Zieldatum.
    svensson_df : pd.DataFrame
        Index = DatetimeIndex, Columns = [Beta0..Tau2].
    method : {'ffill', 'exact'}, default 'ffill'
        - 'ffill': letzter Handelstag ≤ date (Wochenenden/Feiertage werden
          rückwärts auf den letzten verfügbaren Tag gemappt)
        - 'exact': KeyError, wenn date kein Handelstag ist.
    """
    target = pd.Timestamp(date)

    if method == "exact":
        if target not in svensson_df.index:
            raise KeyError(
                f"{target.date()} ist kein Handelstag im Svensson-DataFrame"
            )
        row = svensson_df.loc[target]
    elif method == "ffill":
        idx = svensson_df.index[svensson_df.index <= target]
        if len(idx) == 0:
            raise KeyError(f"Kein Handelstag ≤ {target.date()} verfügbar")
        row = svensson_df.loc[idx[-1]]
    else:
        raise ValueError(f"method must be 'exact' or 'ffill', got {method!r}")

    return params_from_row(row)


# ----------------------------------------------------------------------
# 5. Stress-Operationen
# ----------------------------------------------------------------------
def shift_curve(
    params: SvenssonParams,
    *,
    dlevel: float = 0.0,
    dslope: float = 0.0,
    dcurv1: float = 0.0,
    dcurv2: float = 0.0,
) -> SvenssonParams:
    """Additive Shocks auf die β-Koeffizienten (τ-Parameter unverändert).

    Alle Shocks in den gleichen Einheiten wie die β-Parameter,
    d.h. Prozentpunkte.

    Parameters
    ----------
    dlevel : Δβ₀ — Parallelverschiebung der Kurve.
    dslope : Δβ₁ — wirkt stark am kurzen Ende.
    dcurv1 : Δβ₂ — erster Buckel (um τ₁).
    dcurv2 : Δβ₃ — zweiter Buckel (um τ₂).
    """
    return SvenssonParams(
        beta0=params.beta0 + dlevel,
        beta1=params.beta1 + dslope,
        beta2=params.beta2 + dcurv1,
        beta3=params.beta3 + dcurv2,
        tau1=params.tau1,
        tau2=params.tau2,
    )


# ----------------------------------------------------------------------
# 6. Komfort-Wrapper
# ----------------------------------------------------------------------
def curve_grid(
    params: SvenssonParams,
    maturities: np.ndarray | None = None,
    *,
    as_decimal: bool = False,
) -> pd.Series:
    """Komplette Kurve über ein Maturity-Grid.

    Default-Grid: 0.25y bis 30y in 0.25y-Schritten (120 Punkte).
    Ergebnis als pd.Series, Index = Maturities, Name = 'zero_rate'.
    """
    if maturities is None:
        maturities = np.arange(0.25, 30.25, 0.25)
    rates = zero_rate(np.asarray(maturities, dtype=float), params,
                      as_decimal=as_decimal)
    return pd.Series(rates, index=maturities, name="zero_rate")


# ============================================================================
#  VALIDATION  ·  python backend/svensson.py
# ============================================================================

def _spot_test_2020_01_02() -> bool:
    """Hardcoded gegen Excel-Werte aus Zeile 11 (2020-01-02)."""
    p = SvenssonParams(
        beta0=0.70701, beta1=-1.39679, beta2=-2.36462, beta3=0.76271,
        tau1=2.94907, tau2=1.64556,
    )
    expected = {
        1: -0.6422517390283221,
        2: -0.6089951320302713,
        5: -0.46798572087893675,
        10: -0.16179140290837649,
        20: 0.2184379475561728,
        30: 0.3791952302576727,
    }
    print("  [1] Spot-Tests (2020-01-02):")
    all_ok = True
    for T, exp in expected.items():
        got = zero_rate(T, p)
        err = abs(got - exp)
        status = "OK" if err < 1e-12 else "FAIL"
        print(f"      T={T:>2}y  computed={got:+.10f}  "
              f"expected={exp:+.10f}  err={err:.1e}  {status}")
        all_ok &= err < 1e-12
    return all_ok


def _math_limits() -> bool:
    """τ→0 Limit + Discount-Identity + as_decimal-Konsistenz."""
    p = SvenssonParams(
        beta0=2.0, beta1=-1.0, beta2=0.5, beta3=0.0, tau1=2.0, tau2=4.0
    )
    print("  [2] Mathematische Eigenschaften:")

    # τ→0: y(0+) = β₀ + β₁ (φ(0)=1, e^0=1)
    y0 = zero_rate(1e-6, p)
    exp0 = p.beta0 + p.beta1
    err0 = abs(y0 - exp0)
    ok0 = err0 < 1e-4
    print(f"      τ→0: y(1e-6)={y0:+.6f}  expected={exp0:+.6f}  "
          f"err={err0:.1e}  {'OK' if ok0 else 'FAIL'}")

    # Discount-Identity: D(T) = exp(-r/100·T)
    df = discount_factor(5.0, p)
    exp_df = np.exp(-zero_rate(5.0, p) / 100.0 * 5.0)
    err_df = abs(df - exp_df)
    ok_df = err_df < 1e-12
    print(f"      D(5)={df:.8f}  exp(-r·T)={exp_df:.8f}  "
          f"err={err_df:.1e}  {'OK' if ok_df else 'FAIL'}")

    # as_decimal-Konsistenz: r_pct == r_dec * 100
    r_pct = zero_rate(10.0, p, as_decimal=False)
    r_dec = zero_rate(10.0, p, as_decimal=True)
    err_dec = abs(r_pct - r_dec * 100.0)
    ok_dec = err_dec < 1e-12
    print(f"      r_pct={r_pct:+.8f}  r_dec*100={r_dec*100:+.8f}  "
          f"err={err_dec:.1e}  {'OK' if ok_dec else 'FAIL'}")

    return ok0 and ok_df and ok_dec


def _shift_test() -> bool:
    """Level-Shift = parallel rate shift; andere Shifts ändern β korrekt."""
    p = SvenssonParams(
        beta0=1.0, beta1=-0.5, beta2=0.2, beta3=0.1, tau1=2.0, tau2=4.0
    )
    print("  [3] Curve Shifts:")

    # Level-Shift +1pp → alle Zinsen +1pp
    p_shifted = shift_curve(p, dlevel=1.0)
    diffs = (zero_rate(np.array([1, 5, 10, 30]), p_shifted)
             - zero_rate(np.array([1, 5, 10, 30]), p))
    err_lvl = float(np.abs(diffs - 1.0).max())
    ok_lvl = err_lvl < 1e-12
    print(f"      dlevel=+1pp → max|Δr − 1pp| = {err_lvl:.1e}  "
          f"{'OK' if ok_lvl else 'FAIL'}")

    # Original-Objekt darf nicht mutiert werden
    ok_immut = (p.beta0 == 1.0)
    print(f"      Immutability of original: {'OK' if ok_immut else 'FAIL'}")

    return ok_lvl and ok_immut


def _historical_lookup_test() -> bool:
    """historical_curve mit synthetischem DataFrame, Wochenend-ffill prüfen."""
    print("  [4] historical_curve (ffill):")
    df = pd.DataFrame(
        {
            "Beta0": [1.0, 2.0],
            "Beta1": [0.0, 0.0],
            "Beta2": [0.0, 0.0],
            "Beta3": [0.0, 0.0],
            "Tau1":  [1.0, 1.0],
            "Tau2":  [2.0, 2.0],
        },
        index=pd.to_datetime(["2024-01-05", "2024-01-08"]),  # Fr / Mo
    )
    # Samstag 2024-01-06 → muss auf Freitag 01-05 mappen
    p_sat = historical_curve("2024-01-06", df)
    ok_ffill = (p_sat.beta0 == 1.0)
    print(f"      Sa→Fr ffill: β₀={p_sat.beta0}  "
          f"{'OK' if ok_ffill else 'FAIL'}")

    # Exact-Mode auf Samstag → KeyError
    try:
        historical_curve("2024-01-06", df, method="exact")
        ok_exact = False
    except KeyError:
        ok_exact = True
    print(f"      method='exact' auf Sa → KeyError: "
          f"{'OK' if ok_exact else 'FAIL'}")

    return ok_ffill and ok_exact


def _full_excel_sweep() -> bool:
    """Vollständiger Vergleich gegen Sheet 'Risk Free Rates'."""
    print("  [5] Vollständiger Excel-Sweep:")
    try:
        import openpyxl
    except ImportError:
        print("      [SKIP] openpyxl nicht installiert.")
        return True

    xlsx = _ROOT / "Risk_free_rates_stresstest.xlsx"
    if not xlsx.exists():
        print(f"      [SKIP] {xlsx.name} nicht gefunden.")
        return True

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Risk Free Rates"]

    # Maturity-Header in Zeile 9 ab Spalte 9
    maturities = []
    for c in range(9, ws.max_column + 1):
        v = ws.cell(row=9, column=c).value
        if v is None:
            break
        maturities.append(float(v))
    maturities = np.array(maturities)

    n_rows = 0
    n_pts = 0
    max_err = 0.0
    for r in range(11, ws.max_row + 1):
        if ws.cell(row=r, column=2).value in (None, ""):
            continue
        try:
            params = SvenssonParams(
                beta0=float(ws.cell(row=r, column=3).value),
                beta1=float(ws.cell(row=r, column=4).value),
                beta2=float(ws.cell(row=r, column=5).value),
                beta3=float(ws.cell(row=r, column=6).value),
                tau1=float(ws.cell(row=r, column=7).value),
                tau2=float(ws.cell(row=r, column=8).value),
            )
        except (TypeError, ValueError):
            continue

        computed = zero_rate(maturities, params)
        excel = []
        for j in range(len(maturities)):
            v = ws.cell(row=r, column=9 + j).value
            excel.append(np.nan if isinstance(v, str) else v)
        excel = np.array(excel, dtype=float)

        mask = ~np.isnan(excel)
        if mask.any():
            errs = np.abs(computed[mask] - excel[mask])
            max_err = max(max_err, float(errs.max()))
            n_pts += int(mask.sum())
        n_rows += 1

    ok = max_err < 1e-9
    print(f"      Verglichen: {n_rows:,} Tage × bis zu {len(maturities)} "
          f"Maturities = {n_pts:,} Punkte")
    print(f"      Max abs error: {max_err:.2e} %  "
          f"{'OK' if ok else 'FAIL'}")
    return ok


def _round_trip_with_real_data() -> bool:
    """params_from_row Round-Trip mit echter Bundesbank-CSV (falls Cache da)."""
    print("  [6] Round-Trip mit echtem Svensson-DataFrame:")
    try:
        from config import CACHE_DIR
    except ImportError:
        print("      [SKIP] config nicht importierbar.")
        return True
    cache = CACHE_DIR / "svensson_params.parquet"
    if not cache.exists():
        print(f"      [SKIP] {cache.name} nicht im Cache "
              f"(führe tools/data_fetch/04_fetch_bundesbank_svensson.py aus).")
        return True

    df = pd.read_parquet(cache)
    last_date = df.index[-1]
    p_via_lookup = historical_curve(last_date, df)
    p_via_row = params_from_row(df.iloc[-1])
    ok = p_via_lookup == p_via_row
    print(f"      historical_curve == params_from_row für {last_date.date()}: "
          f"{'OK' if ok else 'FAIL'}")
    return ok


if __name__ == "__main__":
    import sys as _sys
    if _sys.platform == "win32":
        _sys.stdout.reconfigure(encoding="utf-8")
    print("=" * 70)
    print(" svensson.py · Validation")
    print("=" * 70)

    results = [
        _spot_test_2020_01_02(),
        _math_limits(),
        _shift_test(),
        _historical_lookup_test(),
        _full_excel_sweep(),
        _round_trip_with_real_data(),
    ]

    print("=" * 70)
    if all(results):
        print(f" [PASS]  Alle {len(results)} Test-Blöcke bestanden.")
    else:
        n_fail = sum(1 for r in results if not r)
        print(f" [FAIL]  {n_fail}/{len(results)} Test-Blöcke fehlgeschlagen.")
    print("=" * 70)
